import os
import csv
import pandas as pd
from openpyxl import load_workbook, Workbook
import xlrd

# (Brad) Need to change filepath
# Divides XLSX tabs
# keyword - XLSX original downloaded file
# outpath - sheet name + '.xlsx' (saved to multiple files)
def divide_tabs(xlsxoriginal):
    wb = load_workbook(filename=xlsxoriginal)

    for sheet in wb.worksheets:
        new_wb = Workbook()
        ws = new_wb.active
        for row_data in sheet.iter_rows():
            for row_cell in row_data:
                ws[row_cell.coordinate].value = row_cell.value

        new_wb.save('/Users/katherinenewcomb/Desktop/TestingRepo/{0}.xlsx'.format(sheet.title))

# xlsxpath - specific xls sheet name (from divided tabs)
# sheetname - usually 'Sheet 1', name of tab in original file
# csvpath - converted xls tab to csv
def csv_from_excel(xlsxpath,sheetname,csvpath):
    wb = xlrd.open_workbook(xlsxpath)
    sh = wb.sheet_by_name(sheetname)
    your_csv_file = open(csvpath, 'w')
    wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

    for rownum in range(sh.nrows):
        wr.writerow(sh.row_values(rownum))

    your_csv_file.close()

# Removes PH from GeoID
# filepath - converted xlsx to csv file or original csv file
# col - column name of GeoID (where 'PH0000000' is)
# outpath - new csv after edits (v1)
def csv_editor(filepath,col,outpath):
    df = pd.read_csv(filepath)
    saved_column = df[col]
    # print(saved_column)

    df[col] = df[col].str[2:-3]
    print(df[col])
    # saved_column = saved_column.rename(columns={"mun_code": "place_name"})

    df.to_csv(outpath, sep=',')

# Removes First Row (bad contents)
# filepath - csv file path
# colname - column name of first row
# identifier - content of first cell
# outpath - mew csv after edits
def remove_row(filepath,colname,identifier,outpath):
    df = pd.read_csv(filepath, index_col = colname)
    df.drop([identifier], inplace = True)
    df.to_csv(outpath, sep=',')

# Merges csv files
# file1 - converted dbf to csv file
# file2 - new csv (downloaded) data to be merged
# geoID1 - column name from file1
# geoID2 - column name from file2
# outpath - new csv after edits
def merge_files(file1, file2, geoID1, geoID2, outpath):
    df = pd.read_csv(file1)
    df2 = pd.read_csv(file2)
    df_merge_difkey = pd.merge(df2, df, how = 'outer', left_on=geoID1, right_on=geoID2)
    print(df_merge_difkey)

    df_merge_difkey.to_csv(outpath, sep=',')

# (Brad) Need to change filepath
def connect_xlsx_files(original_filename,xlsxoriginal,xlsx_divided_tab,sheetname,xlsx_to_csv,colname_csv_edit,edited_csv_outpath,gadmFile,gadm_colname,merged_outpath):
    filepath = r'/Users/katherinenewcomb/Desktop/TestingRepo/censusfiles/'+(original_filename)
    if filepath.endswith('.XLSX'):
        divide_tabs(xlsxoriginal)
        csv_from_excel(xlsx_divided_tab,sheetname,xlsx_to_csv)
        csv_editor(xlsx_to_csv,colname_csv_edit,edited_csv_outpath)
        merge_files(gadmFile,edited_csv_outpath,colname_csv_edit,gadm_colname,merged_outpath)
        print(filepath)
    elif filepath.endswith('.xlsx'):
        divide_tabs(keyword,xlsx_outpath)
        csv_from_excel(xlsxpath,sheetname,csvpath)
        csv_editor(removed_row_outpath,colname_csv_edit,edited_csv_outpath)
        merge_files(gadmFile,edited_csv_outpath,colname_csv_edit,gadm_colname,merged_outpath)
        print(filepath)
    else:
        print('failed')

def connect_csv_files(original_filename, colname_removeRow, identifier, removed_row_outpath, colname_csv_edit, edited_csv_outpath, gadmFile, gadm_colname, merged_outpath):
    filepath = r'/Users/katherinenewcomb/Desktop/TestingRepo/Philippines/ConnectFiles/'+(original_filename)
    if filepath.endswith('.csv'):
        remove_row(filepath,colname_removeRow,identifier,removed_row_outpath)
        csv_editor(removed_row_outpath,colname_csv_edit,edited_csv_outpath)
        merge_files(gadmFile,edited_csv_outpath,colname_csv_edit,gadm_colname,merged_outpath)
        print('works')
    elif filepath.endswith('.CSV'):
        remove_row(filepath,colname_removeRow,identifier,removed_row_outpath)
        csv_editor(removed_row_outpath,colname_csv_edit,edited_csv_outpath)
        merge_files(gadmFile,edited_csv_outpath,colname_csv_edit,gadm_colname,merged_outpath)
        print('working')
    else:
        print('failed')
